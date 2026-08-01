"""파견 근로자 명단 엑셀 업로드/내보내기.

실무에서 파견 근로자 명단·급여대장은 대부분 엑셀로 관리되므로,
폼 입력 대신 일괄 업로드/다운로드를 지원한다.
"""
from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook, load_workbook

from app.engines import rule_checker

DISPATCH_WORKER_IMPORT_HEADERS = ["이름", "직종", "파견시작일"]
DISPATCH_WORKER_EXPORT_HEADERS = [
    "이름",
    "직종",
    "파견시작일",
    "만료예정일",
    "D-Day",
    "상태",
]


def parse_dispatch_workers_excel(file_bytes: bytes) -> list[dict]:
    """엑셀 파일을 파싱해 파견 근로자 레코드 목록을 반환한다.

    1행은 헤더(이름/직종/파견시작일)로 간주하고 2행부터 데이터로 읽는다.
    """
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active

    records: list[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or all(cell in (None, "") for cell in row):
            continue
        name, position, contract_start_date = row[0], row[1], row[2]
        if isinstance(contract_start_date, datetime):
            contract_start_date = contract_start_date.date()
        elif isinstance(contract_start_date, str):
            contract_start_date = date.fromisoformat(contract_start_date)
        records.append(
            {
                "name": str(name).strip(),
                "position": str(position).strip() if position else "",
                "contract_start_date": contract_start_date,
            }
        )
    return records


def export_dispatch_workers_excel(workers: list[dict]) -> bytes:
    """파견 근로자 목록 + 계산된 만료 리스크를 엑셀로 내보낸다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "파견근로자현황"
    sheet.append(DISPATCH_WORKER_EXPORT_HEADERS)

    for worker in workers:
        expiration = rule_checker.check_dispatch_expiration(worker["contract_start_date"])
        sheet.append(
            [
                worker["name"],
                worker.get("position", ""),
                worker["contract_start_date"].isoformat(),
                expiration["limit_date"].isoformat(),
                expiration["d_day"],
                expiration["status"].value,
            ]
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
