import io
from datetime import date

from openpyxl import Workbook, load_workbook

from app.engines import excel_io, rule_checker


def _build_sample_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(excel_io.DISPATCH_WORKER_IMPORT_HEADERS)
    sheet.append(["홍길동", "생산직", "2025-01-01"])
    sheet.append(["김파견", "사무직", date(2024, 6, 1)])
    sheet.append([None, None, None])  # 빈 행은 무시되어야 함
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_dispatch_workers_excel_reads_rows_and_skips_blank():
    records = excel_io.parse_dispatch_workers_excel(_build_sample_workbook())
    assert len(records) == 2
    assert records[0] == {
        "name": "홍길동",
        "position": "생산직",
        "contract_start_date": date(2025, 1, 1),
    }
    assert records[1]["contract_start_date"] == date(2024, 6, 1)


def test_export_dispatch_workers_excel_includes_risk_columns():
    start = date(2025, 1, 1)
    workers = [{"name": "홍길동", "position": "생산직", "contract_start_date": start}]
    content = excel_io.export_dispatch_workers_excel(workers)

    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    assert header == excel_io.DISPATCH_WORKER_EXPORT_HEADERS

    # d_day/상태는 실행 시점(today) 기준으로 계산되므로, 프로덕션 코드와
    # 동일한 함수로 기대값을 계산해 비교한다 (하드코딩하면 매일 깨지는 테스트가 됨).
    expected = rule_checker.check_dispatch_expiration(start)
    row = [cell.value for cell in sheet[2]]
    assert row[0] == "홍길동"
    assert row[1] == "생산직"
    assert row[2] == start.isoformat()
    assert row[3] == expected["limit_date"].isoformat()
    assert row[4] == expected["d_day"]
    assert row[5] == expected["status"].value
